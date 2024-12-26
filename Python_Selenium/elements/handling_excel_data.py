import openpyxl
from openpyxl import load_workbook
import os
import json

excel_file=os.path.join(os.path.dirname(__file__),"..","test_data","DEMOQA.xlsx")
workboox=load_workbook(excel_file)
sheet=workboox['DEMOQA']

json_data={
  "firstName": "Rahul",
  "lastName": "Jaykar",
  "userEmail": "rjaykar@example.com",
  "age": 30,
  "salary": 75000,
  "department": "Engineering"
}

def reading_excel_data():
    total_row=sheet.max_row
    total_col=sheet.max_column
    for r in range(1,total_row+1):
        for c in range(1,total_col+1):
            print(sheet.cell(r,c).value,end="  |  ")
    print("TC_01 : To Read data in Sheet using loop is Passed.")

def reading_data_through_iterable():
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            print(cell.value, end="  ")
        print()
    print("TC_02 : To Read data in Sheet using iterable is Passed.")

def add_new_data_in_excel():
    next_row=sheet.max_row+1
    col_index=1
    for key,value in json_data.items():
        sheet.cell(next_row,col_index,value)
        col_index +=1
    workboox.save(excel_file)
    print("TC_03 : To Add data in Sheet using loop is Passed.")
    reading_data_through_iterable()
    
def add_new_data_using_iter():
    next_row=sheet.max_row+1
    col_index=1
    iterable=iter(json_data.values())
    while True:
        try:
            value=next(iterable)
            sheet.cell(row=next_row,column=col_index,value=value)
            col_index +=1
        except StopIteration :
            break
    workboox.save(excel_file)  
    print("TC_04 : To Add data in Sheet using iterable is Passed.")
    reading_data_through_iterable()      
    
def delete_row_data():
    last_row=sheet.max_row
    sheet.delete_rows(last_row)
    workboox.save(excel_file)
    reading_data_through_iterable()
    print("TC_05 : To delete Row from excel is Passed.")
    
def creating_new_sheet():
    workboox.create_sheet("NewSheet", 2)
    workboox.save(excel_file)
    print("TC_06 : To Create New sheet in excel is Passed.")
    
def delete_sheet():
    sheet_to_remove="NewSheet"
    if sheet_to_remove in workboox.sheetnames:
        sheet=workboox[sheet_to_remove]
        workboox.remove(sheet)
        workboox.save(excel_file)
        print("TC_07 : To remove Sheet from excel is Passed.")
    else:
        print("TC_07 : To remove Sheet from excel is Failed.")

def create_json_object_from_excel_data():
    data=[]
    """fetch headers contain in first row of table"""
    headers=[cell.value for cell in sheet[1]]
    """Now iterate all data from 2nd row and mapped it with header by converting it to dict"""
    for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row  ,values_only=True):
        raw_data=dict(zip(headers,row))
        data.append(raw_data)
    new_json_data=json.dumps(data, indent=4)
    print(new_json_data)
   
def update_excel_sheet_data():
    """Update Salary of Sachin and Pratik to 1500000 & 1500000"""
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1,max_col=sheet.max_column):
        firstname=row[0].value
        lastname=row[1].value
        
        if firstname=="Sachin" and lastname=="Sherki":
            row[4].value=1500000
            #row[4].value=1300000
        elif firstname=="Pratik" and lastname=="Yewale":
            row[4].value=1500000
            #row[4].value=1300000
    workboox.save(excel_file)
    
    
# reading_excel_data()
# reading_data_through_iterable()
# add_new_data_in_excel()
# delete_row_data()
# add_new_data_using_iter()
# delete_row_data()
# creating_new_sheet()
# delete_sheet()
# create_json_object_from_excel_data()
#update_excel_sheet_data()
