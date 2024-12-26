from selenium.webdriver.common.by import By
from elements import base
import time
import os
from openpyxl import load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

headers_xpath="//div[@class='rt-thead -header']//div[@class='rt-resizable-header-content']"
cell_xpath="//div[@class='rt-td']"
row_xpath="//div[@class='rt-tr-group']"
username="Cierra Vega"
excel_file=os.path.join(os.path.dirname(__file__),'..','test_data','DEMOQA.xlsx')


def locate_web_table():
    driver=base.launch_browser()
    element=driver.find_element(By.XPATH,"//h5[text()='Elements']")
    base.execute_js_script(driver, element)
    web_ele=driver.find_element(By.XPATH, "//span[text()='Web Tables']")
    base.execute_js_script(driver, web_ele)
    return driver

def test_case_01_fetch_and_print_all_table_data():
    driver=locate_web_table()
    rows=driver.find_elements(By.XPATH, row_xpath)
    
    for row_index, row in enumerate(rows,start=1):
        cells=row.find_elements(By.XPATH,f"{row_xpath}[{row_index}]{cell_xpath}")
        if cells[0].text.strip() !="":
            for cell_index,cell in enumerate(cells,start=1):
                if cell.text.strip() !="":    
                    print(cell.text.strip(),end=" | ") 
                else:
                    continue  
            print()

def test_case_02_fetch_specific_suer_data():
    user_name=username.split(" ")
    driver=locate_web_table()
    rows=driver.find_elements(By.XPATH, row_xpath)
    
    for row_index, row in enumerate(rows,start=1):
        cells=row.find_elements(By.XPATH,f"{row_xpath}[{row_index}]{cell_xpath}")
        if cells[0].text.strip().lower() ==user_name[0].lower() and cells[1].text.strip().lower()==user_name[1].lower():
            for cell_index,cell in enumerate(cells,start=1):
                print(cell.text.strip(),end=" | ")
            


def read_excel_data(excel_file,sheet_name):
    workbook=load_workbook(excel_file)
    sheet=workbook[sheet_name]
    data=[]
    headers=[cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2,values_only=True):
        data.append(dict(zip(headers,row)))
    return data
    
def enter_data_submit_form(driver,form_data):
    driver.find_element(By.ID, "addNewRecordButton").click()
    form_keys=["firstName","lastName","userEmail","age","salary","department"]
    input_elements=[]
    for key in form_keys:
        input_elements.append(WebDriverWait(driver,5).until(EC.visibility_of_element_located((By.ID,key))))
    for field in input_elements:
        field_name=field.get_attribute("ID")
        if  field_name in form_data:
            field.send_keys(form_data[field_name])
    driver.find_element(By.ID,"submit").click()

def test_case_03_add_user_from_excelsheet():
    form_data=read_excel_data(excel_file, "DEMOQA")
    driver=locate_web_table()
    for data in form_data:
        enter_data_submit_form(driver, data)
        time.sleep(2)
    time.sleep(3)
    
test_case_01_fetch_and_print_all_table_data()
test_case_02_fetch_specific_suer_data()
test_case_03_add_user_from_excelsheet()
    
    
    
    
    